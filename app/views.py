from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib import messages
from dateutil.relativedelta import relativedelta
from django.db.models import Q

from decimal import Decimal
from .models import Parcela, INCCIndex, ConfiguracaoCalculo
from .forms import ParcelaForm, INCCIndexForm
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.views.decorators.http import require_POST
from .utils.incc_base import carregar_incc_no_banco

def base(request):
    return render(request, 'base.html')


def _to_decimal(value):
    if value is None or value == '':
        return None
    try:
        return Decimal(str(value).replace(',', '.'))
    except Exception:
        return None


def filtrar_parcelas(queryset, params):
    nome = (params.get('nome') or '').strip()
    vencimento_de = (params.get('vencimento_de') or '').strip()
    vencimento_ate = (params.get('vencimento_ate') or '').strip()
    pagamento_de = (params.get('pagamento_de') or '').strip()
    pagamento_ate = (params.get('pagamento_ate') or '').strip()
    atraso_min = (params.get('atraso_min') or '').strip()
    atraso_max = (params.get('atraso_max') or '').strip()
    valor_min = (params.get('valor_min') or '').strip()
    valor_max = (params.get('valor_max') or '').strip()

    if nome:
        queryset = queryset.filter(Q(nome__icontains=nome))
    if vencimento_de:
        queryset = queryset.filter(data_vencimento__gte=vencimento_de)
    if vencimento_ate:
        queryset = queryset.filter(data_vencimento__lte=vencimento_ate)
    if pagamento_de:
        queryset = queryset.filter(data_pagamento__gte=pagamento_de)
    if pagamento_ate:
        queryset = queryset.filter(data_pagamento__lte=pagamento_ate)
    if atraso_min.isdigit():
        queryset = queryset.filter(dias_atraso__gte=int(atraso_min))
    if atraso_max.isdigit():
        queryset = queryset.filter(dias_atraso__lte=int(atraso_max))

    valor_min_decimal = _to_decimal(valor_min)
    valor_max_decimal = _to_decimal(valor_max)
    if valor_min_decimal is not None:
        queryset = queryset.filter(valor_total__gte=valor_min_decimal)
    if valor_max_decimal is not None:
        queryset = queryset.filter(valor_total__lte=valor_max_decimal)

    filtros = {
        'nome': nome,
        'vencimento_de': vencimento_de,
        'vencimento_ate': vencimento_ate,
        'pagamento_de': pagamento_de,
        'pagamento_ate': pagamento_ate,
        'atraso_min': atraso_min,
        'atraso_max': atraso_max,
        'valor_min': valor_min,
        'valor_max': valor_max,
    }
    return queryset, filtros

def incc_index_form(request):
    if request.method == 'POST':
        form = INCCIndexForm(request.POST)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, 'Índice INCC salvo com sucesso.')
                return redirect('incc_index')
            except ValueError as e:
                messages.error(request, str(e))
    else:
        form = INCCIndexForm()

    incc_indices = INCCIndex.objects.all().order_by('-mes_ano')
    return render(request, 'incc/incc_index_form.html', {'form': form, 'incc_indices': incc_indices})


# Função para calcular o percentual acumulado do INCC entre duas datas
def calcular_incc_acumulado(data_inicio, data_fim):
    """Retorna o percentual acumulado do INCC do mês inicial até o mês anterior ao pagamento."""
    if data_fim < data_inicio:
        return Decimal('0')

    acumulado = Decimal('0.00')
    atual = data_inicio.replace(day=1)
    limite = data_fim.replace(day=1) - relativedelta(months=1)

    while atual <= limite:
        try:
            indice = INCCIndex.objects.get(mes_ano=atual)
            acumulado += indice.percentual
        except INCCIndex.DoesNotExist:
            pass

        atual += relativedelta(months=1)

    return acumulado

# Função para calcular a parcela
def calcular_parcela(request):
    if request.method == 'POST':
        form = ParcelaForm(request.POST)
        if form.is_valid():
            parcela = form.save(commit=False)
            venc = parcela.data_vencimento
            pagto = parcela.data_pagamento

            if venc and pagto:
                if venc > pagto:
                    form.add_error('data_pagamento', 'Data de pagamento não pode ser anterior ao vencimento.')
                    return render(request, 'parcela/form.html', {'form': form})

            if pagto:
                configuracao = ConfiguracaoCalculo.obter_configuracao()
                multa_percentual = Decimal(str(configuracao.multa_percentual))
                juros_percentual_mensal = Decimal(str(configuracao.juros_percentual_mensal))
                taxa_boleto_configurada = Decimal(str(configuracao.taxa_boleto))

                # Tenta buscar o INCC do mês de pagamento. Se não existir, tenta o do mês anterior.
                mes_ref = pagto.replace(day=1)
                if not INCCIndex.objects.filter(mes_ano=mes_ref).exists():
                    mes_ref = mes_ref - relativedelta(months=1)
                    if not INCCIndex.objects.filter(mes_ano=mes_ref).exists():
                        form.add_error('data_pagamento', 'Não existe índice de INCC para o mês de pagamento ou mês anterior.')
                        return render(request, 'parcela/form.html', {'form': form})

                dias_atraso = (pagto - venc).days
                dias_atraso = max(dias_atraso, 0)
                #Multa Opcional
                multa = Decimal('0.00')  # Valor padrão
                if parcela.aplicar_multa:
                    if dias_atraso > 0:
                        multa = parcela.valor_original * multa_percentual
                # Juros proporcionais (1% ao mês = 0.03333% ao dia, aproximado por 1/30)
                if parcela.aplicar_juros:
                    juros = parcela.valor_original * juros_percentual_mensal * Decimal(dias_atraso) / Decimal('30')
                else:
                    juros = Decimal('0.00')
                # Correção opcional pelo INCC
                if parcela.aplicar_incc:
                    percentual_incc = calcular_incc_acumulado(venc, pagto) / Decimal('100')
                    correcao_incc = parcela.valor_original * percentual_incc
                else:
                    correcao_incc = Decimal('0.00')

                # Taxa de boleto configurada
                taxa_boleto = taxa_boleto_configurada

                # Valor total
                total = parcela.valor_original + multa + juros + correcao_incc + taxa_boleto

                # Preenchendo campos
                parcela.dias_atraso = dias_atraso
                parcela.multa = multa
                parcela.juros_mora = juros
                parcela.correcao_incc = correcao_incc
                parcela.taxa_boleto = taxa_boleto
                parcela.valor_total = total

                parcela.save()
                return redirect('parcela_detalhe', parcela_id=parcela.id)
    else:
        form = ParcelaForm()

    return render(request, 'parcela/form.html', {'form': form})

def parcela_detalhe(request, parcela_id):
    parcela = get_object_or_404(Parcela, id=parcela_id)
    return render(request, 'parcela/detail.html', {'parcela': parcela})

def parcela_list(request):
    parcelas = Parcela.objects.all().order_by('-data_vencimento', '-id')
    parcelas, filtros = filtrar_parcelas(parcelas, request.GET)
    incc_indices = INCCIndex.objects.all().order_by('-mes_ano')
    form = INCCIndexForm()
    return render(request, 'parcela/list.html', {
        'parcelas': parcelas,
        'incc_indices': incc_indices,
        'incc_form': form,
        'filtros': filtros,
    })

def parcela_delete(request, parcela_id):
    parcela = get_object_or_404(Parcela, id=parcela_id)
    if request.method == 'POST':
        parcela.delete()
        return redirect('parcela_list')
    return render(request, 'parcela/list.html', {'parcela': parcela})

def excluir_varias_parcelas(request):
    if request.method == 'POST':
        ids = request.POST.getlist('parcelas_selecionadas')
        if ids:
            Parcela.objects.filter(id__in=ids).delete()
            messages.success(request, f"{len(ids)} parcela(s) excluída(s) com sucesso.")
        else:
            messages.warning(request, "Nenhuma parcela foi selecionada.")
    return redirect('parcela_list')
def gerar_excel(request):
    # Buscar dados das parcelas
    parcelas, _ = filtrar_parcelas(Parcela.objects.all(), request.GET)

    # Criar DataFrame com os dados
    data = []
    for p in parcelas:
        data.append({
            'Data Vencimento': p.data_vencimento.strftime('%d/%m/%Y'),
            'Data Pagamento': p.data_pagamento.strftime('%d/%m/%Y'),
            'Valor Original': float(p.valor_original),
            'Valor Total': float(p.valor_total),
        })

    df = pd.DataFrame(data, columns=['Data Vencimento', 'Data Pagamento', 'Valor Original', 'Valor Total'])

    if df.empty:
        df.loc[len(df.index)] = ['Sem parcelas cadastradas', '-', 0.0, 0.0]
    else:
        # Cálculo dos totais
        total_original = df['Valor Original'].sum()
        total_total = df['Valor Total'].sum()

        # Adicionar linhas de totais
        df.loc[len(df.index)] = ['', '', 'Valor Original:', total_original]
        df.loc[len(df.index)] = ['', '', 'Valor Atualizado:', total_total]

    # Criar um arquivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="parcelas.xlsx"'

    # Gravar no arquivo
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Parcelas', index=False)

        # Estilização com openpyxl
        ws = writer.sheets['Parcelas']

        # Estilo para cabeçalhos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        center_align = Alignment(horizontal='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col_num, column_title in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border

        # Estilo para dados e ajuste de largura
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.border = border
                if isinstance(cell.value, float):
                    cell.number_format = '#,##0.00'

        # Ajustar largura das colunas automaticamente
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

    return response

@require_POST
def alimentar_incc(request):
    # Carregar os dados do INCC no banco de dados
    try:
        carregar_incc_no_banco()  # Chama o código para carregar os dados
        messages.success(request, "Dados do INCC alimentados com sucesso!")
    except Exception as e:
        return HttpResponse(f"Erro ao alimentar os dados INCC: {str(e)}")

    return redirect('incc_index')  # Redireciona de volta para a lista de INCC